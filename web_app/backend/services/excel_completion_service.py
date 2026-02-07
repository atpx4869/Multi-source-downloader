"""
Excel 标准号补全服务
功能：将不完整的标准号补全为完整格式（添加年份）并获取标准名称
"""
import os
import time
import uuid
import threading
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd
from datetime import datetime

from web_app.excel_standard_processor import StandardProcessor


class ExcelCompletionTask:
    """Excel 补全任务"""
    
    def __init__(self, task_id: str):
        self.task_id = task_id
        self.status = 'pending'  # pending, processing, completed, failed
        self.progress = 0
        self.total = 0
        self.current_row = 0
        self.success_count = 0
        self.fail_count = 0
        self.logs: List[str] = []
        self.result_file: Optional[str] = None
        self.error: Optional[str] = None
        self.start_time = time.time()
        self.end_time: Optional[float] = None
    
    def add_log(self, message: str):
        """添加日志"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_entry = f"[{timestamp}] {message}"
        self.logs.append(log_entry)
        # 只保留最近100条日志
        if len(self.logs) > 100:
            self.logs = self.logs[-100:]
    
    @property
    def elapsed_time(self) -> int:
        """已用时间（秒）"""
        if self.end_time:
            return int(self.end_time - self.start_time)
        return int(time.time() - self.start_time)
    
    def to_dict(self) -> dict:
        """转换为字典"""
        return {
            'task_id': self.task_id,
            'status': self.status,
            'progress': self.progress,
            'total': self.total,
            'current_row': self.current_row,
            'success_count': self.success_count,
            'fail_count': self.fail_count,
            'logs': self.logs,
            'error': self.error,
            'elapsed_time': self.elapsed_time,
            'result_file': self.result_file
        }


class ExcelCompletionService:
    """Excel 标准号补全服务"""
    
    def __init__(self):
        self.tasks: Dict[str, ExcelCompletionTask] = {}
        self.upload_dir = Path("uploads")
        self.result_dir = Path("results")
        self.upload_dir.mkdir(exist_ok=True)
        self.result_dir.mkdir(exist_ok=True)
        self.processor = StandardProcessor()
    
    def create_task(self) -> str:
        """创建新任务"""
        task_id = str(uuid.uuid4())
        task = ExcelCompletionTask(task_id)
        self.tasks[task_id] = task
        return task_id
    
    def get_task(self, task_id: str) -> Optional[ExcelCompletionTask]:
        """获取任务"""
        return self.tasks.get(task_id)
    
    def process_file(self, task_id: str, file_path: str, std_column: str = None):
        """处理文件（异步）"""
        task = self.tasks.get(task_id)
        if not task:
            return
        
        # 在后台线程处理
        thread = threading.Thread(
            target=self._process_in_background,
            args=(task_id, file_path, std_column),
            daemon=True
        )
        thread.start()
    
    def _read_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """读取文件"""
        file_ext = Path(file_path).suffix.lower()
        
        try:
            if file_ext in ['.xlsx', '.xls']:
                return pd.read_excel(file_path)
            elif file_ext == '.csv':
                return pd.read_csv(file_path, encoding='utf-8-sig')
            elif file_ext == '.txt':
                return pd.read_csv(file_path, sep='\t', encoding='utf-8-sig')
            else:
                return None
        except Exception as e:
            print(f"读取文件失败: {e}")
            return None
    
    def _identify_std_column(self, df: pd.DataFrame, std_column: str = None) -> Optional[str]:
        """识别标准号列"""
        if std_column and std_column in df.columns:
            return std_column
        
        # 自动识别：查找包含"标准"或"号"的列
        for col in df.columns:
            if '标准' in str(col) or '号' in str(col):
                return col
        
        # 默认使用第一列
        return df.columns[0] if len(df.columns) > 0 else None
    
    def _process_in_background(
        self,
        task_id: str,
        file_path: str,
        std_column: str = None
    ):
        """后台处理"""
        task = self.tasks.get(task_id)
        if not task:
            return
        
        try:
            task.status = 'processing'
            task.add_log("开始处理文件...")
            
            # 1. 读取文件
            df = self._read_file(file_path)
            if df is None or df.empty:
                task.status = 'failed'
                task.error = "文件读取失败或为空"
                task.add_log("❌ 文件读取失败或为空")
                return
            
            task.add_log(f"✓ 读取到 {len(df)} 行数据")
            
            # 2. 识别标准号列
            std_col = self._identify_std_column(df, std_column)
            if not std_col:
                task.status = 'failed'
                task.error = "无法识别标准号列"
                task.add_log("❌ 无法识别标准号列")
                return
            
            task.add_log(f"✓ 使用列: {std_col}")
            
            # 3. 提取标准号
            std_numbers = df[std_col].tolist()
            task.total = len(std_numbers)
            task.add_log(f"开始处理 {task.total} 个标准号...")
            
            # 4. 处理每个标准号
            results = []
            for idx, std_no in enumerate(std_numbers):
                task.current_row = idx + 1
                
                # 跳过空值
                if pd.isna(std_no) or str(std_no).strip() == '':
                    results.append({
                        '原始标准号': '',
                        '补全标准号': '',
                        '标准名称': '',
                        '状态': '空值'
                    })
                    task.fail_count += 1
                    continue
                
                try:
                    # 调用 StandardProcessor 处理
                    full_std_no, name, status = self.processor.process_standard(str(std_no))
                    
                    results.append({
                        '原始标准号': str(std_no),
                        '补全标准号': full_std_no,
                        '标准名称': name,
                        '状态': status
                    })
                    
                    if '成功' in status:
                        task.success_count += 1
                    else:
                        task.fail_count += 1
                    
                except Exception as e:
                    results.append({
                        '原始标准号': str(std_no),
                        '补全标准号': '',
                        '标准名称': '',
                        '状态': f'错误: {str(e)[:30]}'
                    })
                    task.fail_count += 1
                
                # 更新进度
                task.progress = int((idx + 1) / task.total * 100)
                
                # 每10个记录一次日志
                if (idx + 1) % 10 == 0:
                    task.add_log(f"已处理 {idx + 1}/{task.total} 个标准号")
            
            # 5. 保存结果
            result_df = pd.DataFrame(results)
            result_filename = f"{task_id}_completion_result.xlsx"
            result_path = self.result_dir / result_filename
            
            # 使用 openpyxl 保存并格式化
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "补全结果"
            
            # 写入表头
            headers = ['原始标准号', '补全标准号', '标准名称', '状态']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row_idx, row_data in enumerate(results, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data[header]
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 状态列着色
                    if header == '状态':
                        if '成功' in str(value):
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif '空值' in str(value) or '未找到' in str(value):
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif '错误' in str(value) or '失败' in str(value):
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 设置列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=len(results)+1, min_col=1, max_col=4):
                for cell in row:
                    cell.border = thin_border
            
            wb.save(str(result_path))
            
            task.result_file = str(result_path)
            task.status = 'completed'
            task.end_time = time.time()
            task.progress = 100
            task.add_log(f"✓ 处理完成！成功: {task.success_count}, 失败: {task.fail_count}")
            
        except Exception as e:
            task.status = 'failed'
            task.error = str(e)
            task.end_time = time.time()
            task.add_log(f"❌ 处理失败: {str(e)}")
    
    def export_csv(self, task_id: str) -> Optional[str]:
        """导出为 CSV"""
        task = self.tasks.get(task_id)
        if not task or not task.result_file:
            return None
        
        try:
            # 读取 Excel 结果
            df = pd.read_excel(task.result_file)
            
            # 保存为 CSV
            csv_path = Path(task.result_file).with_suffix('.csv')
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            
            return str(csv_path)
        except Exception as e:
            print(f"导出 CSV 失败: {e}")
            return None
