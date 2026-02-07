# -*- coding: utf-8 -*-
"""
标准查新服务
批量查询标准元数据（发布日期、实施日期、状态、替代标准等）
"""
import os
import sys
import re
import uuid
import time
import asyncio
import threading
from pathlib import Path
from typing import Optional, List, Dict, Any, Callable
from datetime import datetime
import pandas as pd

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from core.aggregated_downloader import AggregatedDownloader


class StandardCheckTask:
    """标准查新任务"""
    
    def __init__(self, task_id: str):
        self.task_id = task_id
        self.status = 'pending'  # pending, processing, completed, failed
        self.progress = 0
        self.total = 0
        self.current_row = 0
        self.success_count = 0
        self.fail_count = 0
        self.logs: List[str] = []
        self.result_df: Optional[pd.DataFrame] = None
        self.result_file: Optional[str] = None
        self.error: Optional[str] = None
        self.start_time = time.time()
        self.end_time: Optional[float] = None
    
    def add_log(self, message: str):
        """添加日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.logs.append(f"[{timestamp}] {message}")
        # 保留最近100条日志
        if len(self.logs) > 100:
            self.logs = self.logs[-100:]
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        elapsed = self.end_time - self.start_time if self.end_time else time.time() - self.start_time
        return {
            'task_id': self.task_id,
            'status': self.status,
            'progress': self.progress,
            'total': self.total,
            'current_row': self.current_row,
            'success_count': self.success_count,
            'fail_count': self.fail_count,
            'elapsed_time': round(elapsed, 1),
            'logs': self.logs[-20:],  # 只返回最近20条日志
            'result_file': self.result_file,
            'error': self.error
        }


class StandardCheckService:
    """标准查新服务"""
    
    def __init__(self, result_dir: str = None):
        self.result_dir = Path(result_dir) if result_dir else Path(__file__).parent.parent.parent / 'results'
        self.result_dir.mkdir(parents=True, exist_ok=True)
        self.tasks: Dict[str, StandardCheckTask] = {}
    
    def create_task(self) -> str:
        """创建新任务"""
        task_id = str(uuid.uuid4())
        self.tasks[task_id] = StandardCheckTask(task_id)
        return task_id
    
    def get_task(self, task_id: str) -> Optional[StandardCheckTask]:
        """获取任务"""
        return self.tasks.get(task_id)
    
    def process_file(
        self,
        task_id: str,
        file_path: str,
        sources: List[str] = None,
        std_column: str = None
    ):
        """
        处理文件（在后台线程中执行）
        
        Args:
            task_id: 任务ID
            file_path: 上传的文件路径
            sources: 启用的数据源列表
            std_column: 标准号列名（可选，自动识别）
        """
        thread = threading.Thread(
            target=self._process_in_background,
            args=(task_id, file_path, sources, std_column)
        )
        thread.daemon = True
        thread.start()
    
    def _process_in_background(
        self,
        task_id: str,
        file_path: str,
        sources: List[str] = None,
        std_column: str = None
    ):
        """后台处理线程"""
        task = self.tasks.get(task_id)
        if not task:
            return
        
        try:
            task.status = 'processing'
            task.add_log("开始处理文件...")
            
            # 1. 读取文件
            df = self._read_file(file_path)
            if df is None or df.empty:
                task.status = 'failed'
                task.error = "文件读取失败或为空"
                task.add_log("❌ 文件读取失败或为空")
                return
            
            task.add_log(f"✓ 读取到 {len(df)} 行数据")
            
            # 2. 识别标准号列
            std_col = self._identify_std_column(df, std_column)
            if not std_col:
                task.status = 'failed'
                task.error = "未找到标准号列"
                task.add_log("❌ 未找到标准号列")
                return
            
            task.add_log(f"✓ 识别标准号列: {std_col}")
            
            # 3. 提取标准号
            std_list = df[std_col].dropna().astype(str).tolist()
            task.total = len(std_list)
            task.add_log(f"✓ 找到 {len(std_list)} 个标准号")
            
            # 4. 初始化下载器
            enabled_sources = sources or ['ZBY']
            task.add_log(f"使用数据源: {', '.join(enabled_sources)}")
            downloader = AggregatedDownloader(enable_sources=enabled_sources, output_dir="downloads")
            
            # 5. 批量查询
            results = []
            for i, std_no in enumerate(std_list, 1):
                task.current_row = i
                task.progress = int((i / len(std_list)) * 100)
                task.add_log(f"查询中 ({i}/{len(std_list)}): {std_no}")
                
                try:
                    result_data = self._query_single_standard(downloader, std_no)
                    results.append(result_data)
                    
                    if result_data.get('标准名称') and result_data['标准名称'] != '未找到':
                        task.success_count += 1
                    else:
                        task.fail_count += 1
                        
                except Exception as e:
                    task.fail_count += 1
                    results.append({
                        '原始标准号': std_no,
                        '规范标准号': '',
                        '标准名称': f'查询失败: {str(e)}',
                        '发布日期': '',
                        '实施日期': '',
                        '状态': '',
                        '替代标准': '',
                        '替代实施日期': '',
                        '替代标准名称': '',
                        '是否变更': ''
                    })
            
            # 6. 保存结果
            task.result_df = pd.DataFrame(results)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_filename = f"标准查新结果_{timestamp}.xlsx"
            result_path = self.result_dir / result_filename
            
            self._export_excel(task.result_df, str(result_path))
            task.result_file = result_filename
            
            task.status = 'completed'
            task.progress = 100
            task.end_time = time.time()
            task.add_log(f"✓ 处理完成！成功 {task.success_count}，失败 {task.fail_count}")
            task.add_log(f"✓ 结果已保存: {result_filename}")
            
        except Exception as e:
            task.status = 'failed'
            task.error = str(e)
            task.end_time = time.time()
            task.add_log(f"❌ 处理失败: {str(e)}")
    
    def _query_single_standard(self, downloader: AggregatedDownloader, std_no: str) -> Dict[str, Any]:
        """查询单个标准"""
        has_year = bool(re.search(r'-\d{4}$', std_no.strip()))
        
        def extract_year(r):
            match = re.search(r'-(\d{4})$', r.std_no)
            return int(match.group(1)) if match else 0
        
        # 搜索标准
        search_results = downloader.search(std_no, limit=10)
        
        # 如果带年代号，额外搜索不带年代号的版本
        all_versions = []
        if has_year and search_results:
            base_std_no = re.sub(r'-\d{4}$', '', std_no.strip())
            all_versions = downloader.search(base_std_no, limit=10)
        
        if search_results:
            # 智能选择
            if not has_year and len(search_results) > 1:
                current_results = [r for r in search_results if r.status and "现行" in r.status]
                if current_results:
                    current_results.sort(key=extract_year, reverse=True)
                    result = current_results[0]
                else:
                    search_results.sort(key=extract_year, reverse=True)
                    result = search_results[0]
            else:
                result = search_results[0]
            
            # 智能判断替代标准和变更状态
            replace_std_text = result.replace_std or ''
            replace_implement = ''
            replace_name = ''
            is_changed = ''
            
            if replace_std_text.strip():
                is_changed = '变更'
            
            # 查找更新版本
            if not replace_std_text.strip():
                if not all_versions:
                    base_std_no = re.sub(r'-\d{4}$', '', result.std_no.strip())
                    all_versions = downloader.search(base_std_no, limit=10)
                
                if all_versions and len(all_versions) > 1:
                    current_year = extract_year(result)
                    base_std_no = re.sub(r'-\d{4}$', '', result.std_no.strip())
                    
                    same_base_versions = [
                        r for r in all_versions 
                        if re.sub(r'-\d{4}$', '', r.std_no.strip()) == base_std_no
                    ]
                    
                    newer_versions = [r for r in same_base_versions if extract_year(r) > current_year]
                    if newer_versions:
                        newer_versions.sort(key=extract_year, reverse=True)
                        newest = newer_versions[0]
                        replace_std_text = newest.std_no
                        replace_implement = newest.implement or ''
                        replace_name = newest.name or ''
                        is_changed = '变更'
            
            if result.status and ('废止' in result.status or '即将废止' in result.status):
                is_changed = '变更'
            
            return {
                '原始标准号': std_no,
                '规范标准号': result.std_no,
                '标准名称': result.name,
                '发布日期': result.publish or '',
                '实施日期': result.implement or '',
                '状态': result.status or '',
                '替代标准': replace_std_text,
                '替代实施日期': replace_implement,
                '替代标准名称': replace_name,
                '是否变更': is_changed
            }
        else:
            return {
                '原始标准号': std_no,
                '规范标准号': '',
                '标准名称': '未找到',
                '发布日期': '',
                '实施日期': '',
                '状态': '',
                '替代标准': '',
                '替代实施日期': '',
                '替代标准名称': '',
                '是否变更': ''
            }
    
    def _read_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """读取文件"""
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xls']:
                return pd.read_excel(file_path)
            elif file_ext == '.csv':
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        return pd.read_csv(file_path, encoding=encoding)
                    except:
                        continue
                return None
            elif file_ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f if line.strip()]
                return pd.DataFrame({'标准号': lines})
            else:
                return None
        except Exception as e:
            print(f"读取文件失败: {e}")
            return None
    
    def _identify_std_column(self, df: pd.DataFrame, user_column: str = None) -> Optional[str]:
        """识别标准号列"""
        if user_column and user_column in df.columns:
            return user_column
        
        possible_names = ['标准号', 'std_no', '标准编号', '编号', 'number', 'code', '标准代号']
        for col in df.columns:
            if any(name in str(col).lower() for name in possible_names):
                return col
        
        return df.columns[0] if len(df.columns) > 0 else None
    
    def _export_excel(self, df: pd.DataFrame, file_path: str):
        """导出为 Excel（包含格式）"""
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
            
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
            status_active_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            status_active_font = Font(color="155724", bold=False)
            status_obsolete_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            status_obsolete_font = Font(color="721C24", bold=False)
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            change_col_idx = None
            status_col_idx = None
            if max_row > 0:
                for col in range(1, max_col + 1):
                    header_cell = worksheet.cell(row=1, column=col)
                    if header_cell.value == '是否变更':
                        change_col_idx = col
                    elif header_cell.value == '状态':
                        status_col_idx = col
            
            for row in range(1, max_row + 1):
                worksheet.row_dimensions[row].height = 14
                
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                    
                    if row == 1:
                        continue
                    
                    if col == status_col_idx and cell.value:
                        cell_value = str(cell.value)
                        if '现行' in cell_value:
                            cell.fill = status_active_fill
                            cell.font = status_active_font
                        elif '废止' in cell_value:
                            cell.fill = status_obsolete_fill
                            cell.font = status_obsolete_font
                    
                    if col == change_col_idx and cell.value == '变更':
                        cell.fill = light_blue_fill
            
            for col_idx in range(1, max_col + 1):
                max_length = 0
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                
                for row in range(1, max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            for char in str(cell.value):
                                if ord(char) > 127:
                                    cell_length += 1
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            workbook.save(file_path)
        except ImportError:
            pass  # openpyxl格式化不可用，使用基础导出
