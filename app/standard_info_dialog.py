# -*- coding: utf-8 -*-
"""
标准查新对话框
用户上传标准号列表，自动查询发布日期、实施日期、状态等元数据信息
"""
import sys
import threading
from pathlib import Path
from datetime import datetime
import pandas as pd

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.http_aggregated_downloader import HttpAggregatedDownloader as AggregatedDownloader

try:
    from PySide6 import QtCore, QtWidgets, QtGui
    PYSIDE_VER = 6
except ImportError:
    from PySide2 import QtCore, QtWidgets, QtGui
    PYSIDE_VER = 2


class StandardInfoDialog(QtWidgets.QDialog):
    """标准查新对话框"""
    
    def __init__(self, parent=None, parent_settings=None):
        super().__init__(parent)
        self.setWindowTitle("标准查新 - 批量查询元数据")
        self.setGeometry(50, 50, 1400, 800)  # 加宽窗口以容纳更多内容
        self.setModal(True)
        self.parent_settings = parent_settings  # 保存主程序配置
        
        # 设置对话框样式
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
            }
            QGroupBox {
                color: #333333;
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                color: #2196F3;
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
            }
            QTableWidget {
                background-color: white;
                color: #333333;
                gridline-color: #ddd;
                selection-background-color: #2196F3;
                selection-color: white;
            }
            QTableWidget::item {
                color: #333333;
            }
            QHeaderView::section {
                background-color: #e8e8e8;
                color: #333333;
                padding: 5px;
                border: 1px solid #ddd;
                font-weight: bold;
            }
        """)
        
        self.downloader = None
        self.result_df = None
        self.file_path = None
        self.processing = False
        
        self.init_ui()
    
    def init_ui(self):
        """初始化 UI"""
        main_layout = QtWidgets.QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # ========== 1. 文件选择区 ==========
        file_group = QtWidgets.QGroupBox("1. 选择标准号文件")
        file_layout = QtWidgets.QHBoxLayout()
        
        self.label_file = QtWidgets.QLabel("未选择文件")
        self.label_file.setStyleSheet("color: #666666; padding: 5px;")
        
        self.btn_select = QtWidgets.QPushButton("选择文件 (Excel/CSV/TXT)")
        self.btn_select.setMaximumWidth(160)
        self.btn_select.clicked.connect(self.select_file)
        self.btn_select.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        file_layout.addWidget(self.label_file, 1)
        file_layout.addWidget(self.btn_select)
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # ========== 2. 配置区 ==========
        config_group = QtWidgets.QGroupBox("2. 配置查询参数")
        config_layout = QtWidgets.QFormLayout()
        
        # 数据源选择
        source_layout = QtWidgets.QHBoxLayout()
        self.cb_zby = QtWidgets.QCheckBox("ZBY")
        self.cb_by = QtWidgets.QCheckBox("BY")
        self.cb_gbw = QtWidgets.QCheckBox("GBW")
        
        # 刷新按钮
        self.btn_refresh_sources = QtWidgets.QPushButton("刷新")
        self.btn_refresh_sources.setMaximumWidth(60)
        self.btn_refresh_sources.clicked.connect(self.refresh_sources_from_parent)
        self.btn_refresh_sources.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 5px 10px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #455A64;
            }
        """)
        
        # 初始化数据源选择（从主程序配置加载）
        self._init_sources_from_parent()
        
        source_layout.addWidget(self.cb_zby)
        source_layout.addWidget(self.cb_by)
        source_layout.addWidget(self.cb_gbw)
        source_layout.addWidget(self.btn_refresh_sources)
        source_layout.addStretch()
        config_layout.addRow("数据源:", source_layout)
        
        # 列名称配置
        self.input_column = QtWidgets.QLineEdit()
        self.input_column.setPlaceholderText("例如: 标准号, std_no, 标准编号 (留空则自动识别)")
        config_layout.addRow("标准号列名:", self.input_column)
        
        config_group.setLayout(config_layout)
        main_layout.addWidget(config_group)
        
        # ========== 3. 处理按钮 ==========
        btn_layout = QtWidgets.QHBoxLayout()
        
        self.btn_process = QtWidgets.QPushButton("🔍 开始查询")
        self.btn_process.setEnabled(False)
        self.btn_process.clicked.connect(self.process_file)
        self.btn_process.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_process)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)
        
        # ========== 4. 进度显示 ==========
        self.progress = QtWidgets.QProgressBar()
        self.progress.setVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 4px;
                text-align: center;
                background-color: white;
            }
            QProgressBar::chunk {
                background-color: #2196F3;
            }
        """)
        main_layout.addWidget(self.progress)
        
        self.label_status = QtWidgets.QLabel("")
        self.label_status.setStyleSheet("color: #666666; padding: 5px;")
        self.label_status.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addWidget(self.label_status)
        
        # ========== 5. 结果显示表格 ==========
        result_group = QtWidgets.QGroupBox("3. 查询结果")
        result_layout = QtWidgets.QVBoxLayout()
        
        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            '原始标准号', '规范标准号', '标准名称', 
            '发布日期', '实施日期', '状态', 
            '替代标准', '替代实施日期', '替代标准名称',
            '是否变更'
        ])
        self.table.horizontalHeader().setStretchLastSection(False)
        # 设置列宽模式：原始标准号、规范标准号、替代标准自适应内容，标准名称、替代名称拉伸
        self.table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QtWidgets.QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        
        result_layout.addWidget(self.table)
        result_group.setLayout(result_layout)
        main_layout.addWidget(result_group, 1)
        
        # ========== 6. 底部按钮 ==========
        bottom_layout = QtWidgets.QHBoxLayout()
        
        self.btn_export_excel = QtWidgets.QPushButton("导出 Excel")
        self.btn_export_excel.setMaximumWidth(100)
        self.btn_export_excel.setEnabled(False)
        self.btn_export_excel.clicked.connect(self.export_excel)
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.btn_export_csv = QtWidgets.QPushButton("导出 CSV")
        self.btn_export_csv.setMaximumWidth(100)
        self.btn_export_csv.setEnabled(False)
        self.btn_export_csv.clicked.connect(self.export_csv)
        self.btn_export_csv.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.btn_close = QtWidgets.QPushButton("关闭")
        self.btn_close.setMaximumWidth(80)
        self.btn_close.clicked.connect(self.close)
        self.btn_close.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.btn_export_excel)
        bottom_layout.addWidget(self.btn_export_csv)
        bottom_layout.addWidget(self.btn_close)
        
        main_layout.addLayout(bottom_layout)
        self.setLayout(main_layout)
    
    def select_file(self):
        """选择文件"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "选择标准号文件",
            "",
            "所有支持的文件 (*.xlsx *.xls *.csv *.txt);;Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv);;文本文件 (*.txt)"
        )
        
        if file_path:
            self.file_path = file_path
            self.label_file.setText(f"已选择: {Path(file_path).name}")
            self.label_file.setStyleSheet("color: #4CAF50; padding: 5px; font-weight: bold;")
            self.btn_process.setEnabled(True)
    
    def _init_sources_from_parent(self):
        """从主程序配置初始化数据源选择"""
        if self.parent_settings and 'sources' in self.parent_settings:
            parent_sources = self.parent_settings['sources']
            self.cb_zby.setChecked('ZBY' in parent_sources)
            self.cb_by.setChecked('BY' in parent_sources)
            self.cb_gbw.setChecked('GBW' in parent_sources)
        else:
            # 默认只选择ZBY
            self.cb_zby.setChecked(True)
            self.cb_by.setChecked(False)
            self.cb_gbw.setChecked(False)
    
    def refresh_sources_from_parent(self):
        """刷新数据源选择（从主程序配置）"""
        self._init_sources_from_parent()
        QtWidgets.QMessageBox.information(self, "提示", "已同步主程序的数据源配置")
    
    def get_enabled_sources(self):
        """获取启用的数据源"""
        sources = []
        if self.cb_zby.isChecked():
            sources.append("ZBY")
        if self.cb_by.isChecked():
            sources.append("BY")
        if self.cb_gbw.isChecked():
            sources.append("GBW")
        return sources if sources else ["ZBY"]
    
    def process_file(self):
        """处理文件"""
        if self.processing:
            QtWidgets.QMessageBox.warning(self, "提示", "正在处理中，请稍候...")
            return
        
        if not self.file_path:
            QtWidgets.QMessageBox.warning(self, "提示", "请先选择文件！")
            return
        
        self.processing = True
        self.btn_process.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.label_status.setText("正在读取文件...")
        
        # 在后台线程处理
        thread = threading.Thread(target=self._process_in_background)
        thread.daemon = True
        thread.start()
    
    def _process_in_background(self):
        """后台处理线程"""
        try:
            # 1. 读取文件
            self.update_status("正在读取文件...", 10)
            df = self._read_file()
            
            if df is None or df.empty:
                self.update_status("文件读取失败或为空！", 0)
                self.processing = False
                return
            
            # 2. 识别标准号列
            self.update_status("正在识别标准号列...", 20)
            std_column = self._identify_std_column(df)
            
            if not std_column:
                self.update_status("未找到标准号列！", 0)
                self.processing = False
                return
            
            # 3. 提取标准号
            std_list = df[std_column].dropna().astype(str).tolist()
            self.update_status(f"找到 {len(std_list)} 个标准号", 30)
            
            # 4. 初始化下载器
            sources = self.get_enabled_sources()
            self.update_status(f"初始化数据源: {', '.join(sources)}", 40)
            self.downloader = AggregatedDownloader(enable_sources=sources, output_dir="downloads")
            
            # 5. 批量查询
            results = []
            total = len(std_list)
            for i, std_no in enumerate(std_list, 1):
                progress = 40 + int((i / total) * 50)
                self.update_status(f"查询中 ({i}/{total}): {std_no}", progress)
                
                # 搜索标准 - 增加limit以便智能筛选
                import re
                has_year = bool(re.search(r'-\d{4}$', std_no.strip()))
                
                # 提取年份的辅助函数
                def extract_year(r):
                    match = re.search(r'-(\d{4})$', r.std_no)
                    return int(match.group(1)) if match else 0
                
                # 先搜索精确匹配
                search_results = self.downloader.search(std_no, limit=10)
                
                # 如果带年代号，额外搜索不带年代号的版本以查找更新版本
                all_versions = []
                if has_year and search_results:
                    base_std_no = re.sub(r'-\d{4}$', '', std_no.strip())
                    all_versions = self.downloader.search(base_std_no, limit=10)
                
                if search_results:
                    # 智能选择：如果不带年代号，优先选择现行标准
                    
                    # 提取年份的辅助函数
                    def extract_year(r):
                        match = re.search(r'-(\d{4})$', r.std_no)
                        return int(match.group(1)) if match else 0
                    
                    if not has_year and len(search_results) > 1:
                        # 不带年代号，优先选择现行标准
                        current_results = [r for r in search_results if r.status and "现行" in r.status]
                        if current_results:
                            # 如果有多个现行标准，选择年份最新的
                            current_results.sort(key=extract_year, reverse=True)
                            result = current_results[0]
                        else:
                            # 没有现行标准，选择年份最新的
                            search_results.sort(key=extract_year, reverse=True)
                            result = search_results[0]
                    else:
                        # 带年代号或只有一个结果，直接使用
                        result = search_results[0]
                    
                    # 智能判断替代标准和变更状态
                    replace_std_text = result.replace_std or ''
                    replace_implement = ''
                    replace_name = ''
                    is_changed = ''
                    
                    # 情况1：API明确返回了替代标准
                    if replace_std_text.strip():
                        is_changed = '变更'
                    
                    # 情况2和3：尝试通过搜索所有版本来查找替代标准
                    # 无论是否带年代号，都尝试查找更新版本
                    if not replace_std_text.strip():  # 只有在API没有返回替代标准时才查找
                        # 获取所有版本（如果还没获取）
                        if not all_versions:
                            base_std_no = re.sub(r'-\d{4}$', '', result.std_no.strip())
                            all_versions = self.downloader.search(base_std_no, limit=10)
                        
                        if all_versions and len(all_versions) > 1:
                            # 只看基础标准号完全相同的版本
                            current_year = extract_year(result)
                            base_std_no = re.sub(r'-\d{4}$', '', result.std_no.strip())
                            
                            # 过滤出基础标准号完全相同的版本
                            same_base_versions = [
                                r for r in all_versions 
                                if re.sub(r'-\d{4}$', '', r.std_no.strip()) == base_std_no
                            ]
                            
                            # 在同一标准号的版本中查找更新的
                            newer_versions = [r for r in same_base_versions if extract_year(r) > current_year]
                            if newer_versions:
                                # 找到了更新的版本
                                newer_versions.sort(key=extract_year, reverse=True)
                                newest = newer_versions[0]
                                replace_std_text = newest.std_no
                                replace_implement = newest.implement or ''
                                replace_name = newest.name or ''
                                is_changed = '变更'
                    
                    # 如果状态为废止，确保标记为变更（即使没找到替代标准）
                    if result.status and ('废止' in result.status or '即将废止' in result.status):
                        is_changed = '变更'
                    
                    results.append({
                        '原始标准号': std_no,
                        '规范标准号': result.std_no,
                        '标准名称': result.name,
                        '发布日期': result.publish or '',
                        '实施日期': result.implement or '',
                        '状态': result.status or '',
                        '替代标准': replace_std_text,
                        '替代实施日期': replace_implement,
                        '替代标准名称': replace_name,
                        '是否变更': is_changed
                    })
                else:
                    results.append({
                        '原始标准号': std_no,
                        '规范标准号': '',
                        '标准名称': '未找到',
                        '发布日期': '',
                        '实施日期': '',
                        '状态': '',
                        '替代标准': '',
                        '替代实施日期': '',
                        '替代标准名称': '',
                        '是否变更': ''
                    })
            
            # 6. 显示结果
            self.result_df = pd.DataFrame(results)
            self.update_status(f"查询完成！找到 {len(results)} 条结果", 100)
            self._display_results()
            
        except Exception as e:
            self.update_status(f"处理失败: {str(e)}", 0)
            QtWidgets.QMessageBox.critical(self, "错误", f"处理失败:\n{str(e)}")
        
        finally:
            self.processing = False
            self.btn_process.setEnabled(True)
    
    def _read_file(self):
        """读取文件"""
        try:
            file_ext = Path(self.file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xls']:
                return pd.read_excel(self.file_path)
            elif file_ext == '.csv':
                # 尝试不同编码
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        return pd.read_csv(self.file_path, encoding=encoding)
                    except:
                        continue
                return None
            elif file_ext == '.txt':
                # 读取为单列
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f if line.strip()]
                return pd.DataFrame({'标准号': lines})
            else:
                return None
        except Exception as e:
            print(f"读取文件失败: {e}")
            return None
    
    def _identify_std_column(self, df):
        """识别标准号列"""
        # 用户指定的列名
        user_column = self.input_column.text().strip()
        if user_column and user_column in df.columns:
            return user_column
        
        # 自动识别
        possible_names = ['标准号', 'std_no', '标准编号', '编号', 'number', 'code', '标准代号']
        for col in df.columns:
            if any(name in str(col).lower() for name in possible_names):
                return col
        
        # 默认使用第一列
        return df.columns[0] if len(df.columns) > 0 else None
    
    def _display_results(self):
        """显示结果"""
        if self.result_df is None or self.result_df.empty:
            return
        
        self.table.setRowCount(0)
        for idx, row in self.result_df.iterrows():
            row_pos = self.table.rowCount()
            self.table.insertRow(row_pos)
            
            for col_idx, col_name in enumerate(self.result_df.columns):
                value = str(row[col_name])
                item = QtWidgets.QTableWidgetItem(value)
                
                # 状态列颜色
                if col_name == '状态' and value:
                    if '现行' in value or 'active' in value.lower():
                        item.setBackground(QtGui.QColor("#d4edda"))
                        item.setForeground(QtGui.QColor("#155724"))
                    elif '废止' in value or 'supersede' in value.lower():
                        item.setBackground(QtGui.QColor("#f8d7da"))
                        item.setForeground(QtGui.QColor("#721c24"))
                
                # 是否变更列颜色（浅蓝色）
                if col_name == '是否变更' and value == '变更':
                    item.setBackground(QtGui.QColor("#cfe2ff"))  # 浅蓝色
                    item.setForeground(QtGui.QColor("#084298"))  # 深蓝色文字
                
                self.table.setItem(row_pos, col_idx, item)
        
        self.btn_export_excel.setEnabled(True)
        self.btn_export_csv.setEnabled(True)
    
    def update_status(self, message, progress):
        """更新状态（线程安全）"""
        QtCore.QMetaObject.invokeMethod(
            self.label_status,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, message)
        )
        QtCore.QMetaObject.invokeMethod(
            self.progress,
            "setValue",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(int, progress)
        )
    
    def export_excel(self):
        """导出为 Excel（包含格式）"""
        if self.result_df is None or self.result_df.empty:
            QtWidgets.QMessageBox.warning(self, "提示", "暂无结果可导出！")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"标准查新结果_{timestamp}.xlsx"
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "导出 Excel", default_name, "Excel 文件 (*.xlsx)"
        )
        
        if file_path:
            try:
                # 导出到Excel
                self.result_df.to_excel(file_path, index=False, engine='openpyxl')
                
                # 使用openpyxl添加格式
                from openpyxl import load_workbook
                from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
                
                workbook = load_workbook(file_path)
                worksheet = workbook.active
                
                # 定义边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 定义填充色（浅蓝色，与UI一致）
                light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
                
                # 定义状态颜色（与UI保持一致）
                # 现行 - 浅绿色
                status_active_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                status_active_font = Font(color="155724", bold=False)
                # 废止 - 浅红色
                status_obsolete_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                status_obsolete_font = Font(color="721C24", bold=False)
                
                # 遍历所有单元格，添加边框、设置行高和对齐方式
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # 查找"是否变更"和"状态"列的位置
                change_col_idx = None
                status_col_idx = None
                if max_row > 0:
                    for col in range(1, max_col + 1):
                        header_cell = worksheet.cell(row=1, column=col)
                        if header_cell.value == '是否变更':
                            change_col_idx = col
                        elif header_cell.value == '状态':
                            status_col_idx = col

                
                for row in range(1, max_row + 1):
                    # 设置行高为14
                    worksheet.row_dimensions[row].height = 14
                    
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        # 添加边框
                        cell.border = thin_border
                        # 设置对齐方式和自动换行
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        
                        # 跳过表头行
                        if row == 1:
                            continue
                        
                        # 根据状态列的值应用不同的填充色
                        if col == status_col_idx and cell.value:
                            cell_value = str(cell.value)
                            if '现行' in cell_value or 'active' in cell_value.lower():
                                cell.fill = status_active_fill
                                cell.font = status_active_font
                            elif '废止' in cell_value or 'supersede' in cell_value.lower():
                                cell.fill = status_obsolete_fill
                                cell.font = status_obsolete_font
                        
                        # 如果是"是否变更"列且值为"变更"，添加填充色
                        if col == change_col_idx and cell.value == '变更':
                            cell.fill = light_blue_fill
                
                # 自适应列宽
                for col_idx in range(1, max_col + 1):
                    max_length = 0
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    
                    for row in range(1, max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        try:
                            if cell.value:
                                # 计算文本长度（中文按2个字符计算）
                                cell_length = len(str(cell.value))
                                for char in str(cell.value):
                                    if ord(char) > 127:  # 中文字符
                                        cell_length += 1
                                max_length = max(max_length, cell_length)
                        except:
                            pass
                    
                    # 设置列宽（加上一点余量）
                    adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                
                workbook.save(file_path)
                QtWidgets.QMessageBox.information(
                    self, "成功", f"已成功导出到:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")
    
    def export_csv(self):
        """导出为 CSV"""
        if self.result_df is None or self.result_df.empty:
            QtWidgets.QMessageBox.warning(self, "提示", "暂无结果可导出！")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"标准查新结果_{timestamp}.csv"
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "导出 CSV", default_name, "CSV 文件 (*.csv)"
        )
        
        if file_path:
            try:
                self.result_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                QtWidgets.QMessageBox.information(
                    self, "成功", f"已成功导出到:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "错误", f"导出失败:\n{str(e)}")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    dialog = StandardInfoDialog()
    dialog.show()
    sys.exit(app.exec())
